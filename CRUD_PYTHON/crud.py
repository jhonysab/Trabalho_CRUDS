import os

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


class Contato:
    """Representação do contato"""
    def __init__(self, nome: str, telefone: str, profissao: str):
        self.nome = nome
        self.telefone = telefone
        self.profissao = profissao

    def __str__(self):
        return f"Nome: {self.nome}, Telefone: {self.telefone}, Profissão: {self.profissao}"

def carregar_dados(arquivo):
    if not os.path.exists(arquivo):
        return []
    
    wb = load_workbook(arquivo)
    ws = wb.active
    lista_pessoas = []

    for row in ws.iter_rows(min_row=2, values_only=True): 
        nome, telefone, profissao = row
        lista_pessoas.append(Contato(nome, telefone, profissao))
    
    wb.close()
    return lista_pessoas

def salvar_dados(lista_pessoas, arquivo):
    """Salva os contatos da lista no arquivo Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Contatos"

    ws.append(["Nome", "Telefone", "Profissão"])

    for pessoa in lista_pessoas:
        ws.append([pessoa.nome, pessoa.telefone, pessoa.profissao])

    wb.save(arquivo)
    wb.close()

def adicionar_pessoa(lista_pessoas):
    print("\n --- Adicionar Pessoa ---")
    nome = input("Digite o nome: ")
    telefone = input("Digite o telefone: ")
    profissao = input("Digite a profissão: ")

    contato = Contato(nome, telefone, profissao)
    lista_pessoas.append(contato)
    print("Pessoa adicionada com sucesso!")

def listar_pessoas(lista_pessoas):
    print("\n --- Lista de Contatos ---")
    if not lista_pessoas:
        print("Nenhuma pessoa cadastrada.")
    else:
        for i, pessoa in enumerate(lista_pessoas, start=1):
            print(f"{i}. {pessoa}")

def remover_pessoa(lista_pessoas):
    print("\n --- Remover Pessoa ---")
    listar_pessoas(lista_pessoas)
    
    if lista_pessoas:
        try:
            index = int(input("Digite o número da pessoa a ser removida: ")) - 1
            if 0 <= index < len(lista_pessoas):
                removed = lista_pessoas.pop(index)
                print(f"{removed.nome} foi removido da lista.")
            else:
                print("Índice inválido.")
        except ValueError:
            print("Entrada inválida. Por favor, digite um número.")

def modificar_pessoa(lista_pessoas):
    print("\n --- Modificar Pessoa ---")
    listar_pessoas(lista_pessoas)
    
    if lista_pessoas:
        try:
            index = int(input("Digite o número da pessoa a ser modificada: ")) - 1
            if 0 <= index < len(lista_pessoas):
                pessoa = lista_pessoas[index]
                
                print(f"Modificando os dados de {pessoa.nome}:")
                novo_nome = input(f"Novo nome (atual: {pessoa.nome}): ")
                novo_telefone = input(f"Novo telefone (atual: {pessoa.telefone}): ")
                nova_profissao = input(f"Nova profissão (atual: {pessoa.profissao}): ")

                pessoa.nome = novo_nome if novo_nome else pessoa.nome
                pessoa.telefone = novo_telefone if novo_telefone else pessoa.telefone
                pessoa.profissao = nova_profissao if nova_profissao else pessoa.profissao
                
                print(f"{pessoa.nome} foi atualizado com sucesso.")
            else:
                print("Índice inválido.")
        except ValueError:
            print("Entrada inválida. Por favor, digite um número.")

def menu_principal():
    arquivo_excel = "agenda_contatos.xlsx"
    lista_pessoas = carregar_dados(arquivo_excel)

    while True:
        print("\n --- Agenda de Contato ---")
        print("1. Adicionar Pessoa")
        print("2. Listar Pessoas")
        print("3. Remover Pessoa")
        print("4. Modificar Pessoa")
        print("5. Sair")

        opcao = input("Escolha uma opção: ")

        if opcao == "1":
            adicionar_pessoa(lista_pessoas)
            salvar_dados(lista_pessoas, arquivo_excel)
        elif opcao == "2":
            listar_pessoas(lista_pessoas)
        elif opcao == "3":
            remover_pessoa(lista_pessoas)
            salvar_dados(lista_pessoas, arquivo_excel)
        elif opcao == "4":
            modificar_pessoa(lista_pessoas)
            salvar_dados(lista_pessoas, arquivo_excel)
        elif opcao == "5":
            print("Saindo e salvando dados no arquivo...")
            salvar_dados(lista_pessoas, arquivo_excel)
            break
        else:
            print("Opção Inválida! Tente novamente.")

menu_principal()
